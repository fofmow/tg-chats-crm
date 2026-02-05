import io
from collections import defaultdict
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from bot.database.models import PaymentIn, PaymentOut


class ExcelService:
    """Service for generating Excel reports."""
    
    # Styles
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    SUBHEADER_FILL = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
    SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    
    @classmethod
    def generate_period_report(
        cls,
        incoming: list[PaymentIn],
        outgoing: list[PaymentOut],
        period_name: str = "Report",
    ) -> io.BytesIO:
        """Generate comprehensive Excel report for a given period."""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        ws_summary = wb.create_sheet("Summary", 0)
        cls._write_comprehensive_summary(ws_summary, incoming, outgoing, period_name)
        
        ws_daily = wb.create_sheet("By Day")
        cls._write_daily_breakdown(ws_daily, incoming, outgoing)
        
        ws_in = wb.create_sheet("Incoming Payments")
        cls._write_incoming_sheet(ws_in, incoming)
        
        ws_out = wb.create_sheet("Outgoing Payments")
        cls._write_outgoing_sheet(ws_out, outgoing)
        
        ws_teachers = wb.create_sheet("By Teacher")
        cls._write_teachers_breakdown(ws_teachers, incoming)
        
        ws_categories = wb.create_sheet("By Category")
        cls._write_categories_breakdown(ws_categories, outgoing)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    # Keep old method name for backward compatibility
    @classmethod
    def generate_7_days_report(
        cls,
        incoming: list[PaymentIn],
        outgoing: list[PaymentOut],
    ) -> io.BytesIO:
        """Generate comprehensive Excel report for the last 7 days."""
        return cls.generate_period_report(incoming, outgoing, period_name="Last 7 Days")
    
    @classmethod
    def _write_comprehensive_summary(
        cls,
        ws,
        incoming: list[PaymentIn],
        outgoing: list[PaymentOut],
        period_name: str = "Report",
    ):
        """Write comprehensive summary sheet."""
        total_in = sum(p.amount for p in incoming)
        total_out = sum(p.amount for p in outgoing)
        balance = total_in - total_out
        
        # Title
        ws.cell(row=1, column=1, value=f"FINANCIAL REPORT: {period_name.upper()}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")
        
        # Period info
        if incoming or outgoing:
            all_dates = [p.date for p in incoming] + [p.date for p in outgoing]
            start_date = min(all_dates) if all_dates else date.today()
            end_date = max(all_dates) if all_dates else date.today()
            ws.cell(row=2, column=1, value=f"Period: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}")
        else:
            ws.cell(row=2, column=1, value="Period: No data")
        ws.merge_cells("A2:D2")
        
        ws.cell(row=3, column=1, value=f"Generated: {date.today().strftime('%d.%m.%Y')}")
        ws.merge_cells("A3:D3")
        
        row = 5
        
        # Main metrics
        ws.cell(row=row, column=1, value="KEY METRICS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(f"A{row}:D{row}")
        row += 1
        
        # Incoming section
        ws.cell(row=row, column=1, value="INCOMING PAYMENTS (DEBIT)")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = cls.SUCCESS_FILL
        ws.merge_cells(f"A{row}:B{row}")
        row += 1
        
        ws.cell(row=row, column=1, value="Total Amount:")
        ws.cell(row=row, column=2, value=total_in)
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        row += 1
        
        ws.cell(row=row, column=1, value="Transaction Count:")
        ws.cell(row=row, column=2, value=len(incoming))
        row += 1
        
        if incoming:
            avg_in = total_in / len(incoming)
            ws.cell(row=row, column=1, value="Average Payment:")
            ws.cell(row=row, column=2, value=avg_in)
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            row += 1
            
            max_in = max(p.amount for p in incoming)
            min_in = min(p.amount for p in incoming)
            ws.cell(row=row, column=1, value="Max Payment:")
            ws.cell(row=row, column=2, value=max_in)
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            row += 1
            
            ws.cell(row=row, column=1, value="Min Payment:")
            ws.cell(row=row, column=2, value=min_in)
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            row += 1
            
            # By chat type
            ru_payments = [p for p in incoming if p.chat_type == "ru"]
            eng_payments = [p for p in incoming if p.chat_type == "eng"]
            ws.cell(row=row, column=1, value="From RU chat:")
            ws.cell(row=row, column=2, value=f"{len(ru_payments)} pcs. for {sum(p.amount for p in ru_payments):,.2f}")
            row += 1
            ws.cell(row=row, column=1, value="From ENG chat:")
            ws.cell(row=row, column=2, value=f"{len(eng_payments)} pcs. for {sum(p.amount for p in eng_payments):,.2f}")
            row += 1
        
        row += 1
        
        # Outgoing section
        ws.cell(row=row, column=1, value="OUTGOING PAYMENTS (CREDIT)")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = cls.WARNING_FILL
        ws.merge_cells(f"A{row}:B{row}")
        row += 1
        
        ws.cell(row=row, column=1, value="Total Amount:")
        ws.cell(row=row, column=2, value=total_out)
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        row += 1
        
        ws.cell(row=row, column=1, value="Transaction Count:")
        ws.cell(row=row, column=2, value=len(outgoing))
        row += 1
        
        if outgoing:
            avg_out = total_out / len(outgoing)
            ws.cell(row=row, column=1, value="Average Payment:")
            ws.cell(row=row, column=2, value=avg_out)
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            row += 1
            
            max_out = max(p.amount for p in outgoing)
            min_out = min(p.amount for p in outgoing)
            ws.cell(row=row, column=1, value="Max Payment:")
            ws.cell(row=row, column=2, value=max_out)
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            row += 1
            
            ws.cell(row=row, column=1, value="Min Payment:")
            ws.cell(row=row, column=2, value=min_out)
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            row += 1
        
        row += 1
        
        # Balance
        ws.cell(row=row, column=1, value="TOTAL BALANCE")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2, value=balance)
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        if balance >= 0:
            ws.cell(row=row, column=2).fill = cls.SUCCESS_FILL
        else:
            ws.cell(row=row, column=2).fill = cls.WARNING_FILL
        
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30
    
    @classmethod
    def _write_daily_breakdown(
        cls,
        ws,
        incoming: list[PaymentIn],
        outgoing: list[PaymentOut],
    ):
        """Write daily breakdown sheet."""
        headers = ["Date", "Incoming (count)", "Incoming (amount)", 
                   "Outgoing (count)", "Outgoing (amount)", "Day Balance"]
        cls._write_headers(ws, headers)
        
        # Group by date
        in_by_date = defaultdict(list)
        out_by_date = defaultdict(list)
        
        for p in incoming:
            in_by_date[p.date].append(p)
        for p in outgoing:
            out_by_date[p.date].append(p)
        
        # Get all dates in range
        all_dates = set(in_by_date.keys()) | set(out_by_date.keys())
        if not all_dates:
            # Show last 7 days even if empty
            today = date.today()
            all_dates = {today - timedelta(days=i) for i in range(7)}
        
        row = 2
        total_in = 0
        total_out = 0
        
        for d in sorted(all_dates, reverse=True):
            day_in = in_by_date.get(d, [])
            day_out = out_by_date.get(d, [])
            
            sum_in = sum(p.amount for p in day_in)
            sum_out = sum(p.amount for p in day_out)
            day_balance = sum_in - sum_out
            
            total_in += sum_in
            total_out += sum_out
            
            ws.cell(row=row, column=1, value=d.strftime("%d.%m.%Y")).border = cls.BORDER
            ws.cell(row=row, column=2, value=len(day_in)).border = cls.BORDER
            ws.cell(row=row, column=3, value=sum_in).border = cls.BORDER
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=len(day_out)).border = cls.BORDER
            ws.cell(row=row, column=5, value=sum_out).border = cls.BORDER
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6, value=day_balance).border = cls.BORDER
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            
            if day_balance >= 0:
                ws.cell(row=row, column=6).fill = cls.SUCCESS_FILL
            else:
                ws.cell(row=row, column=6).fill = cls.WARNING_FILL
            
            row += 1
        
        # Total row
        row += 1
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=len(incoming)).font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_in).font = Font(bold=True)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=len(outgoing)).font = Font(bold=True)
        ws.cell(row=row, column=5, value=total_out).font = Font(bold=True)
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=total_in - total_out).font = Font(bold=True)
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        
        # Adjust widths
        widths = [12, 18, 18, 18, 18, 15]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    @classmethod
    def _write_teachers_breakdown(cls, ws, incoming: list[PaymentIn]):
        """Write breakdown by teacher."""
        headers = ["Teacher", "Count", "Amount", "Average", "% of Total"]
        cls._write_headers(ws, headers)
        
        if not incoming:
            ws.cell(row=2, column=1, value="No data")
            return
        
        # Group by teacher
        by_teacher = defaultdict(list)
        for p in incoming:
            by_teacher[p.teacher].append(p)
        
        total_amount = sum(p.amount for p in incoming)
        
        # Sort by amount descending
        sorted_teachers = sorted(
            by_teacher.items(),
            key=lambda x: sum(p.amount for p in x[1]),
            reverse=True
        )
        
        row = 2
        for teacher, payments in sorted_teachers:
            total = sum(p.amount for p in payments)
            avg = total / len(payments)
            pct = (total / total_amount * 100) if total_amount > 0 else 0
            
            ws.cell(row=row, column=1, value=teacher).border = cls.BORDER
            ws.cell(row=row, column=2, value=len(payments)).border = cls.BORDER
            ws.cell(row=row, column=3, value=total).border = cls.BORDER
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=avg).border = cls.BORDER
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=f"{pct:.1f}%").border = cls.BORDER
            row += 1
        
        # Total
        row += 1
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=len(incoming)).font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_amount).font = Font(bold=True)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        
        widths = [25, 12, 15, 15, 12]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    @classmethod
    def _write_categories_breakdown(cls, ws, outgoing: list[PaymentOut]):
        """Write breakdown by category."""
        headers = ["Category", "Count", "Amount", "Average", "% of Total"]
        cls._write_headers(ws, headers)
        
        if not outgoing:
            ws.cell(row=2, column=1, value="No data")
            return
        
        # Group by category
        by_category = defaultdict(list)
        for p in outgoing:
            by_category[p.category].append(p)
        
        total_amount = sum(p.amount for p in outgoing)
        
        # Sort by amount descending
        sorted_categories = sorted(
            by_category.items(),
            key=lambda x: sum(p.amount for p in x[1]),
            reverse=True
        )
        
        row = 2
        for category, payments in sorted_categories:
            total = sum(p.amount for p in payments)
            avg = total / len(payments)
            pct = (total / total_amount * 100) if total_amount > 0 else 0
            
            ws.cell(row=row, column=1, value=category).border = cls.BORDER
            ws.cell(row=row, column=2, value=len(payments)).border = cls.BORDER
            ws.cell(row=row, column=3, value=total).border = cls.BORDER
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=avg).border = cls.BORDER
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=f"{pct:.1f}%").border = cls.BORDER
            row += 1
        
        # Total
        row += 1
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=len(outgoing)).font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_amount).font = Font(bold=True)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        
        widths = [25, 12, 15, 15, 12]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    @classmethod
    def _write_incoming_sheet(cls, ws, payments: list[PaymentIn]):
        """Write incoming payments to worksheet."""
        headers = ["#", "Date", "Amount", "Client", "Teacher", "Chat", "Added"]
        cls._write_headers(ws, headers)
        
        for i, payment in enumerate(payments, start=2):
            ws.cell(row=i, column=1, value=i - 1).border = cls.BORDER
            ws.cell(row=i, column=2, value=payment.date.strftime("%d.%m.%Y")).border = cls.BORDER
            ws.cell(row=i, column=3, value=payment.amount).border = cls.BORDER
            ws.cell(row=i, column=3).number_format = '#,##0.00'
            ws.cell(row=i, column=4, value=payment.client).border = cls.BORDER
            ws.cell(row=i, column=5, value=payment.teacher).border = cls.BORDER
            ws.cell(row=i, column=6, value=payment.chat_type.upper()).border = cls.BORDER
            ws.cell(row=i, column=7, value=payment.created_at.strftime("%d.%m.%Y %H:%M")).border = cls.BORDER
        
        # Add total row
        if payments:
            total_row = len(payments) + 2
            ws.cell(row=total_row, column=2, value="TOTAL:").font = Font(bold=True)
            ws.cell(row=total_row, column=3, value=sum(p.amount for p in payments)).font = Font(bold=True)
            ws.cell(row=total_row, column=3).number_format = '#,##0.00'
        
        cls._adjust_column_widths(ws, headers)
    
    @classmethod
    def _write_outgoing_sheet(cls, ws, payments: list[PaymentOut]):
        """Write outgoing payments to worksheet."""
        headers = ["#", "Date", "Amount", "Category", "Recipient", "Added"]
        cls._write_headers(ws, headers)
        
        for i, payment in enumerate(payments, start=2):
            ws.cell(row=i, column=1, value=i - 1).border = cls.BORDER
            ws.cell(row=i, column=2, value=payment.date.strftime("%d.%m.%Y")).border = cls.BORDER
            ws.cell(row=i, column=3, value=payment.amount).border = cls.BORDER
            ws.cell(row=i, column=3).number_format = '#,##0.00'
            ws.cell(row=i, column=4, value=payment.category).border = cls.BORDER
            ws.cell(row=i, column=5, value=payment.recipient).border = cls.BORDER
            ws.cell(row=i, column=6, value=payment.created_at.strftime("%d.%m.%Y %H:%M")).border = cls.BORDER
        
        # Add total row
        if payments:
            total_row = len(payments) + 2
            ws.cell(row=total_row, column=2, value="TOTAL:").font = Font(bold=True)
            ws.cell(row=total_row, column=3, value=sum(p.amount for p in payments)).font = Font(bold=True)
            ws.cell(row=total_row, column=3).number_format = '#,##0.00'
        
        cls._adjust_column_widths(ws, headers)
    
    @classmethod
    def _write_headers(cls, ws, headers: list[str]):
        """Write styled headers to worksheet."""
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.border = cls.BORDER
            cell.alignment = cls.CENTER_ALIGN
    
    @classmethod
    def _adjust_column_widths(cls, ws, headers: list[str]):
        """Adjust column widths based on content."""
        min_widths = {
            "#": 5,
            "Date": 12,
            "Amount": 15,
            "Client": 20,
            "Teacher": 20,
            "Chat": 8,
            "Chat Type": 12,
            "Category": 20,
            "Recipient": 20,
            "Added": 16,
        }
        
        for col, header in enumerate(headers, start=1):
            width = min_widths.get(header, 15)
            ws.column_dimensions[get_column_letter(col)].width = width
